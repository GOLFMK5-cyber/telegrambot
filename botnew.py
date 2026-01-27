from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, KeyboardButton, ReplyKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler, filters,
    ConversationHandler, ContextTypes, CallbackQueryHandler
)
import datetime
import sqlite3
import os
import logging
from logging.handlers import RotatingFileHandler
from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, "residents.db")
REQUESTS_LOG = os.path.join(BASE_DIR, "requests.txt")
EXCEL_FILE = os.path.join(BASE_DIR, "requests.xlsx")

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

file_handler = RotatingFileHandler(
    os.path.join(BASE_DIR, "bot.log"),
    maxBytes=5*1024*1024,
    backupCount=3,
    encoding="utf-8"
)
file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s:%(message)s"))
logger.addHandler(file_handler)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.WARNING)
logger.addHandler(console_handler)

def init_db():
    conn = sqlite3.connect(DB_FILE, timeout=10)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS residents (
            telegram_id INTEGER PRIMARY KEY,
            full_name   TEXT NOT NULL,
            flat        TEXT NOT NULL,
            phone       TEXT UNIQUE,
            created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

def normalize_phone(phone: str) -> str:
    if not phone:
        return ""
    digits = ''.join(c for c in phone if c.isdigit())
    if digits.startswith('0') and len(digits) == 10:
        digits = '38' + digits[1:]
    if len(digits) == 9:
        digits = '38' + digits
    return digits

def get_resident_by_phone(phone: str):
    if not phone:
        return None
    normalized = normalize_phone(phone)
    conn = sqlite3.connect(DB_FILE, timeout=10)
    c = conn.cursor()
    c.execute("SELECT telegram_id, full_name, flat FROM residents WHERE phone=? OR phone=?", (phone, normalized))
    row = c.fetchone()
    conn.close()
    if row:
        return {"telegram_id": row[0], "full_name": row[1], "flat": row[2]}
    return None

def save_resident(telegram_id: int, full_name: str, flat: str, phone: str):
    normalized_phone = normalize_phone(phone)
    flat_normalized = flat.upper().strip()
    conn = sqlite3.connect(DB_FILE, timeout=10)
    c = conn.cursor()
    c.execute('''
        INSERT OR REPLACE INTO residents (telegram_id, full_name, flat, phone)
        VALUES (?, ?, ?, ?)
    ''', (telegram_id, full_name.strip(), flat_normalized, normalized_phone))
    conn.commit()
    conn.close()

def save_request(data):
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([
                "ID заявки", "Тип", "Ім’я", "Квартира", "Телефон",
                "Telegram", "Номер авто", "Гість", "Дата/час", "Тривалість", "Статус", "Час логування"
            ])

        row = [
            data[0], "Авто" if data[1] == "auto" else "Гість",
            data[3], data[4], data[5], data[6],
            data[7], data[8], data[9], data[11], data[10],
            datetime.datetime.utcnow().isoformat()
        ]
        ws.append(row)
        wb.save(EXCEL_FILE)
        logger.info("Request logged to Excel: #%s", data[0])
    except Exception:
        logger.exception("Failed to write request to Excel file")

def load_last_request_id():
    if not os.path.exists(REQUESTS_LOG):
        return 0
    try:
        with open(REQUESTS_LOG, "r", encoding="utf-8") as f:
            lines = f.readlines()
        for line in reversed(lines):
            if line.startswith("Номер заявки: #"):
                try:
                    return int(line.strip().split("#")[1])
                except:
                    continue
        return 0
    except Exception:
        logger.exception("Не вдалося зчитати останній номер заявки")
        return 0

MENU, CHOOSING, AUTO_CAR, AUTO_DATE, AUTO_TIME_CHOICE, AUTO_DURATION, \
GUEST_GUESTNAME, GUEST_DATE, GUEST_TIME_CHOICE, GUEST_DURATION, CONTACT, REG_NAME, REG_FLAT = range(13)

SECURITY_CHAT_ID = -1003886695943
ALLOWED_GROUP_ID = -1002485481604
REQUEST_COUNTER = 0
async def is_user_allowed(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    user_id = update.effective_user.id
    try:
        member = await context.bot.get_chat_member(ALLOWED_GROUP_ID, user_id)
        return member.status in ("member", "administrator", "creator")
    except Exception:
        return False

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await is_user_allowed(update, context):
        await update.message.reply_text("⛔ У вас немає доступу до цього бота.")
        return ConversationHandler.END
    keyboard = [[KeyboardButton("📱 Поділитись номером телефону", request_contact=True)]]
    await update.message.reply_text(
        "Ласкаво просимо! Поділіться номером телефону:",
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    )
    return CONTACT

async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = update.message.contact.phone_number if update.message.contact else None
    context.user_data["phone"] = phone
    resident = get_resident_by_phone(phone)
    if resident:
        context.user_data["name"] = resident["full_name"]
        context.user_data["flat"] = resident["flat"]
        await update.message.reply_text(f"Вітаю, {resident['full_name']}! 🏠 Квартира {resident['flat']} підтягнуто.")
    else:
        await update.message.reply_text("Номер не знайдено. Введіть Ваше 👤 Ім’я та Прізвище:")
        return REG_NAME
    await show_main_menu(update.effective_user.id, context)
    return CHOOSING

async def reg_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["name"] = update.message.text.strip()
    await update.message.reply_text("Введіть 🏠 номер квартири:")
    return REG_FLAT

async def reg_flat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["flat"] = text
    save_resident(update.effective_user.id, context.user_data["name"], text, context.user_data["phone"])
    await show_main_menu(update.effective_user.id, context)
    return CHOOSING

async def show_main_menu(user_id, context):
    await context.bot.send_message(
        chat_id=user_id,
        text="Оберіть тип пропуску:",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🚗 Пропуск АВТО", callback_data="auto")],
            [InlineKeyboardButton("👤 Пропуск Гостя", callback_data="guest")],
            [InlineKeyboardButton("📞 Номери телефонів охорони", callback_data="security_numbers")],
            [InlineKeyboardButton("❌ Відмінити", callback_data="end")]
        ])
    )

async def menu_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "end":
        await context.bot.send_message(query.from_user.id, "❌ Діалог відмінено. Почніть заново: /start")
        return ConversationHandler.END
    if query.data in ("back_to_choice", "new_request"):
        await show_main_menu(query.from_user.id, context)
        return CHOOSING
    if query.data == "security_numbers":
        await context.bot.send_message(
            chat_id=query.from_user.id,
            text="📞 Номери телефонів охорони:\n"
                 "🔒 +38 (050) 788-42-14 — Охорона в першій секції\n"
                 "🔒 +38 (095) 384-33-56 — Охорона в другій секції"
        )
        await show_main_menu(query.from_user.id, context)
        return CHOOSING
    if query.data == "auto":
        await context.bot.send_message(query.from_user.id, "Введіть 🚗 номер авто:")
        return AUTO_CAR
    if query.data == "guest":
        await context.bot.send_message(query.from_user.id, "Введіть 👤 Ім’я та Прізвище гостя:")
        return GUEST_GUESTNAME
async def auto_car(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["auto_car"] = update.message.text.strip()
    keyboard = [
        [InlineKeyboardButton("🚗 Авто стоїть у воріт", callback_data="auto_now")],
        [InlineKeyboardButton("📅 Протягом дня", callback_data="auto_today")],
        [InlineKeyboardButton("➡️ На завтра", callback_data="auto_tomorrow")],
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_choice"),
         InlineKeyboardButton("❌ Відмінити", callback_data="end")]
    ]
    await update.message.reply_text("Оберіть дату/час заїзду:", reply_markup=InlineKeyboardMarkup(keyboard))
    return AUTO_DATE

async def auto_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "auto_now":
        context.user_data["auto_date"] = "🚗 Стоїть у воріт"
        return await ask_duration(query, context, "auto")
    elif query.data in ["auto_today", "auto_tomorrow"]:
        context.user_data["auto_day"] = "Сьогодні" if query.data == "auto_today" else "Завтра"
        keyboard = [
            [InlineKeyboardButton("🌅 Перша половина дня", callback_data="auto_firsthalf")],
            [InlineKeyboardButton("🌆 Друга половина дня", callback_data="auto_secondhalf")],
            [InlineKeyboardButton("⏰ Конкретний час", callback_data="auto_exact")]
        ]
        await query.edit_message_text("Оберіть час:", reply_markup=InlineKeyboardMarkup(keyboard))
        return AUTO_TIME_CHOICE

async def auto_time_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        if query.data == "auto_firsthalf":
            context.user_data["auto_date"] = f"{context.user_data['auto_day']} — 🌅 перша половина"
            return await ask_duration(query, context, "auto")
        elif query.data == "auto_secondhalf":
            context.user_data["auto_date"] = f"{context.user_data['auto_day']} — 🌆 друга половина"
            return await ask_duration(query, context, "auto")
        elif query.data == "auto_exact":
            await query.edit_message_text("Введіть ⏰ конкретний час (наприклад 14:30):")
            return AUTO_TIME_CHOICE
    elif update.message:
        context.user_data["auto_date"] = f"{context.user_data.get('auto_day','')} — ⏰ {update.message.text.strip()}"
        return await ask_duration(update, context, "auto")

async def guest_guestname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["guest_guestname"] = update.message.text.strip()
    keyboard = [
        [InlineKeyboardButton("🚶 Гість стоїть у калітки", callback_data="guest_now")],
        [InlineKeyboardButton("📅 Протягом дня", callback_data="guest_today")],
        [InlineKeyboardButton("➡️ На завтра", callback_data="guest_tomorrow")],
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_choice"),
         InlineKeyboardButton("❌ Відмінити", callback_data="end")]
    ]
    await update.message.reply_text("Оберіть дату/час візиту:", reply_markup=InlineKeyboardMarkup(keyboard))
    return GUEST_DATE

async def guest_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "guest_now":
        context.user_data["guest_date"] = "🚶 Стоїть у калітки"
        return await ask_duration(query, context, "guest")
    elif query.data in ["guest_today", "guest_tomorrow"]:
        context.user_data["guest_day"] = "Сьогодні" if query.data == "guest_today" else "Завтра"
        keyboard = [
            [InlineKeyboardButton("🌅 Перша половина дня", callback_data="guest_firsthalf")],
            [InlineKeyboardButton("🌆 Друга половина дня", callback_data="guest_secondhalf")],
            [InlineKeyboardButton("⏰ Конкретний час", callback_data="guest_exact")]
        ]
        await query.edit_message_text("Оберіть час:", reply_markup=InlineKeyboardMarkup(keyboard))
        return GUEST_TIME_CHOICE

async def guest_time_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        if query.data == "guest_firsthalf":
            context.user_data["guest_date"] = f"{context.user_data['guest_day']} — 🌅 перша половина"
            return await ask_duration(query, context, "guest")
        elif query.data == "guest_secondhalf":
            context.user_data["guest_date"] = f"{context.user_data['guest_day']} — 🌆 друга половина"
            return await ask_duration(query, context, "guest")
        elif query.data == "guest_exact":
            await query.edit_message_text("Введіть ⏰ конкретний час (наприклад 18:45):")
            return GUEST_TIME_CHOICE
    elif update.message:
        context.user_data["guest_date"] = f"{context.user_data.get('guest_day','')} — ⏰ {update.message.text.strip()}"
        return await ask_duration(update, context, "guest")

# Вибір тривалості перебування
async def ask_duration(query_or_update, context: ContextTypes.DEFAULT_TYPE, mode: str):
    keyboard = [
        [InlineKeyboardButton("⏱ До години", callback_data=f"{mode}_dur1")],
        [InlineKeyboardButton("⏱ Від 1–2 годин", callback_data=f"{mode}_dur2")],
        [InlineKeyboardButton("⏱ Від 2–4 годин", callback_data=f"{mode}_dur3")]
    ]
    if hasattr(query_or_update, "edit_message_text"):
        await query_or_update.edit_message_text("Оберіть ⏱ час перебування:", reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await query_or_update.message.reply_text("Оберіть ⏱ час перебування:", reply_markup=InlineKeyboardMarkup(keyboard))
    return AUTO_DURATION if mode == "auto" else GUEST_DURATION

async def auto_duration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    durations = {"auto_dur1": "⏱ До години", "auto_dur2": "⏱ 1–2 години", "auto_dur3": "⏱ 2–4 години"}
    context.user_data["duration"] = durations.get(query.data, "")
    return await finalize_request(query, context, "auto")

async def guest_duration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    durations = {"guest_dur1": "⏱ До години", "guest_dur2": "⏱ 1–2 години", "guest_dur3": "⏱ 2–4 години"}
    context.user_data["duration"] = durations.get(query.data, "")
    return await finalize_request(query, context, "guest")
async def finalize_request(query, context: ContextTypes.DEFAULT_TYPE, mode: str):
    global REQUEST_COUNTER
    REQUEST_COUNTER += 1
    req_id = REQUEST_COUNTER

    data = [
        req_id, mode, None,
        context.user_data.get("name"), context.user_data.get("flat"),
        context.user_data.get("phone"), query.from_user.username,
        context.user_data.get("auto_car") if mode == "auto" else None,
        context.user_data.get("guest_guestname") if mode == "guest" else None,
        context.user_data.get("auto_date") if mode == "auto" else context.user_data.get("guest_date"),
        "Очікує", context.user_data.get("duration")
    ]
    save_request(data)

    # Формування заявки
    summary = f"✅ Заявка #{req_id}\n"
    if mode == "auto":
        summary += f"🚗 Номер авто: {data[7]}\n"
    else:
        summary += f"👤 Гість: {data[8]}\n"
    summary += f"📅 Дата/час: {data[9]}\n"
    summary += f"⏱️ Тривалість: {data[11]}\n\n"
    summary += f"📌 ЗАЯВНИК\n"
    summary += f"👤 Ім’я: {data[3]}\n"
    summary += f"🏠 Квартира: {data[4]}\n"
    summary += f"📞 Телефон: {data[5]}\n"
    summary += f"📲 Telegram: @{data[6]}\n"
    summary += f"📌 Статус: {data[10]}"

    # кнопка для охорони
    keyboard_security = [[InlineKeyboardButton("✅ Прийнято", callback_data=f"accepted_{query.from_user.id}_{req_id}")]]
    await context.bot.send_message(SECURITY_CHAT_ID, summary, reply_markup=InlineKeyboardMarkup(keyboard_security))

    # кнопка для заявника
    keyboard_user = [[InlineKeyboardButton("📝 Створити нову заявку", callback_data="new_request")]]
    await context.bot.send_message(chat_id=query.from_user.id, text=summary, reply_markup=InlineKeyboardMarkup(keyboard_user))

    return CHOOSING


async def accepted(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data.startswith("accepted_"):
        parts = query.data.split("_")
        user_id = int(parts[1])
        await context.bot.send_message(user_id, "✅ Ваша заявка прийнята охоронцем.")
        await query.edit_message_text(text=query.message.text + "\n\n✅ Прийнято", reply_markup=None)


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Дію відмінено. Почніть заново: /start")
    return ConversationHandler.END


async def show_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"Ваш Telegram ID: {update.effective_user.id}")


def main():
    global REQUEST_COUNTER
    init_db()
    REQUEST_COUNTER = load_last_request_id()

    token = "8184081641:AAFIZE2A8CQkw5Gzt-J-ZrTBlAwbzWR2qx4"  # встав свій токен від BotFather
    app = Application.builder().token(token).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            CONTACT: [MessageHandler(filters.CONTACT | filters.TEXT, handle_contact)],
            REG_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_name)],
            REG_FLAT: [MessageHandler(filters.TEXT & ~filters.COMMAND, reg_flat)],
            CHOOSING: [CallbackQueryHandler(menu_choice)],
            AUTO_CAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, auto_car)],
            AUTO_DATE: [CallbackQueryHandler(auto_date)],
            AUTO_TIME_CHOICE: [
                CallbackQueryHandler(auto_time_choice),
                MessageHandler(filters.TEXT & ~filters.COMMAND, auto_time_choice)
            ],
            AUTO_DURATION: [CallbackQueryHandler(auto_duration)],
            GUEST_GUESTNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, guest_guestname)],
            GUEST_DATE: [CallbackQueryHandler(guest_date)],
            GUEST_TIME_CHOICE: [
                CallbackQueryHandler(guest_time_choice),
                MessageHandler(filters.TEXT & ~filters.COMMAND, guest_time_choice)
            ],
            GUEST_DURATION: [CallbackQueryHandler(guest_duration)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("id", show_id))
    app.add_handler(CallbackQueryHandler(accepted, pattern="^accepted_"))
    app.add_handler(conv_handler)

    logger.info("🚀 Bot is running...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()